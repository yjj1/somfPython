# -*- coding:'utf8' -*-
#encoding=utf-8

from openpyxl import load_workbook
import sys

def getTargetCell(sheet, value):
    for row in sheet.iter_rows():
        for cell in row:
            # print cell.value
            if cell.value == value:
                return cell.row, cell.column

if __name__ == '__main__':
    reload(sys)
    sys.setdefaultencoding('utf-8')
    fileName = '电子银行查询-电子银行活跃客户量表9670000-1710161103239.xlsx'

    excel = load_workbook(fileName)
    sheetNames = excel.get_sheet_names()
    for sheetName in sheetNames:
        sheet = excel.get_sheet_by_name(sheetName)
        t = getTargetCell(sheet, '企业网银')
        tag1Row = t[0]
        tag1Col = t[1]
        print t
        # t2 = getTargetCell(sheet, '')


        # # print sheet.max_row, sheet.max_column
        # for i in range(1, sheet.max_row):
        #     for j in range(1, sheet.max_column):
        #         print sheet.cell(row=i,column=j).value
        #
        # print sheet.cell(row=5,column=18).value


